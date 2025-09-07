from __future__ import annotations

import re
from datetime import datetime
from functools import wraps
from pathlib import Path
from typing import Iterator

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from starlette.datastructures import URL

from .. import utils

__all__ = [
    'cellstr',
    'cellurl',
    'extract_workbook',
    'extract_worksheet',
    'load_workbook']

logger = utils.get_logger('xlsx')

@wraps(openpyxl.load_workbook)
def load_workbook(file: Path, read_only=True, **kw):
    logger.debug(f'Loading {file.name}')
    return openpyxl.load_workbook(file, read_only=read_only, **kw)

def extract_workbook(file: Path, **kw) -> Iterator[dict[str, str]]:
    ws = load_workbook(file, **kw).worksheets[0]
    return extract_worksheet(ws)

def extract_worksheet(ws: Worksheet, **kw) -> Iterator[dict[str, str]]:
    it = ws.iter_rows(values_only=False, **kw)
    it = (tuple(map(cellstr, cells)) for cells in it)
    headers = next(it)
    it = filter(any, it)
    for values in it:
        yield dict(filter(any, zip(headers, values)))

def cellstr(cell: Cell) -> str:
    value = cell.value
    if value is None:
        value = ''
    elif isinstance(value, datetime):
        value = value.strftime(f'%Y-%m-%d')
    else:
        value = str(value)
    return value

def cellurl(cell: Cell) -> str|None:
    try:
        url = cell.hyperlink and cell.hyperlink.target
    except AttributeError:
        pass
    else:
        if url:
            return str(URL(url))
    value = cellstr(cell)
    value = re.sub(r'=HYPERLINK\(\s*"([^"]+)".*', r'\1', value)
    if value.startswith(('http://', 'https://')):
        return str(URL(value))